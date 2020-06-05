from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

from backpack_solver import backpack_solver


def get_invest_potential(offers, duration):
    investing = []
    for offer in offers:
        profit = (duration - offer["no_income"] - 1) * offer["income"] - offer["start_cost"]
        investing.append((offer["start_cost"], profit, offer["index"]))
    return investing


def get_offers(ws):
    offers = []
    i = 2
    while ws[f"A{i}"].value:
        offers.append({
            "index": i - 1,
            "start_cost": int(ws[f"B{i}"].value),
            "no_income": int(ws[f"C{i}"].value),
            "income": int(ws[f"D{i}"].value),
            "day": 0,
            "start_day": 0
        })
        i += 1
    return offers


def main():
    wb = load_workbook("Предложения.xlsx")
    wb1 = Workbook()
    ws = wb.active
    ws1 = wb1.active
    offers = get_offers(ws)
    offers_length = len(offers)
    start_money, invest_duration = int(ws["F2"].value), int(ws["G2"].value)
    offers_value = get_invest_potential(offers, invest_duration)  # рассчитываем возможную прибыль
    first_result = backpack_solver(offers_value, start_money)  # Решаем задачу о рюкзаке
    use_offers = []
    for i in first_result:  # помечаем данные предложения как использованные
        use_offers.append(offers.pop(i[2] - 1))
    # заполняем статическую информацию
    ws1[f"A{offers_length + 2}"].value = "Инвестирование"
    ws1[f"A{offers_length + 3}"].value = "Сумма:"
    ws1[f"B{offers_length + 3}"].value = start_money
    ws1[f"C{offers_length + 2}"].value = sum(map(lambda x: x["start_cost"], use_offers))
    for i in range(1, offers_length + 1):
        ws1[f"B{i + 1}"].value = i
    current_money = start_money - sum(map(lambda x: x[0], first_result))
    print("FIRST", first_result)
    for i in range(1, invest_duration + 1):
        for j in use_offers:
            if j["day"] >= j["no_income"] + j["start_day"]:
                current_money += j["income"]  # добавляем дневную прибыль к основной
                j["day"] = i
            else:
                j["day"] = i
        offers.sort(key=lambda x: (invest_duration - x["no_income"] - 1) * x["income"] - x["start_cost"],
                    reverse=True)  # сортируем предложения по уменьшению эффективности инвестиций
        offers_value = get_invest_potential(offers, invest_duration - i)
        if offers_value: # если у нас еще остались предложения
            max_invest = max(map(lambda x: x[1], offers_value))
            for offer in enumerate(offers_value):
                if current_money >= offer[1][0] and offer[1][1] > 0 and offer[1][1] == max_invest: #выбираем предложение с максимальной прибылью
                    offers[offer[0]]["start_day"] = i - 1
                    use_offers.append(offers.pop(offer[0]))  # инвестируем в предложение
                    current_money -= offer[1][0]
                    if ws1.cell(row=offers_length + 2, column=i + 2).value:
                        ws1.cell(row=offers_length + 2, column=i + 2).value = offer[1][0] + ws1.cell(row=offers_length + 2,
                                                                                                     column=i + 2).value
                    else:
                        ws1.cell(row=offers_length + 2, column=i + 2).value = offer[1][0]
        print(offers_value, i, invest_duration - i, current_money)
        ws1.cell(row=1, column=i + 2).value = i
        ws1.cell(row=offers_length + 3, column=i + 2).value = current_money  # записываем сумму на текущий день
        for offer in use_offers:
            if offer["day"] > offer["no_income"] + offer["start_day"]:
                ws1.cell(row=offer["index"] + 1, column=i + 2).value = offer["income"]  # записываем дневную прибыть
            else:
                # пока прибыль не началась, помечаем желтым цветом
                ws1.cell(row=offer["index"] + 1, column=i + 2).fill = PatternFill(start_color="FFFFFF00",
                                                                                  fill_type='solid')
    print(current_money)
    wb1.save("Стратегия.xlsx")


if __name__ == "__main__":
    main()
